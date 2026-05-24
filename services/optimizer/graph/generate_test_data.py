"""Synthetic data generator for the ``/graph`` prototype framework.

Produces ``data_experiment/artificial_plans.xlsx`` with five sheets that
mirror the canonical LineWise schemas in
``packages/contracts/module/schemas.py`` at a smaller, fully deterministic
scale:

* ``skus``             — synthetic SKU catalogue (one row per ``SKU``)
* ``demand``           — one ``DemandBucket`` per SKU for one window
* ``line_capability``  — full ``SKU x line`` grid with ``can_produce`` boolean
                         + ``median_speed_uds_per_hour``
* ``node_costs``       — production hours per *feasible* ``(sku_id, line_id)``
                         pair given the current demand
* ``edge_matrix``      — one row per *feasible* directed ``(line_id, sku_from_id,
                         sku_to_id)`` triple: ``total_hours`` (ground truth),
                         ``predicted_hours`` (LightGBM q50), ``q10_hours`` /
                         ``q90_hours`` (uncertainty band), plus the per-segment
                         decomposition (``brand_h``, ``container_h`` …)

Cost model — best-in-class swap target
--------------------------------------

The previous prototype used a hand-rolled additive segment model. This
revision upgrades the changeover predictor to a **gradient-boosted quantile
ensemble** (LightGBM) trained on a synthetic *ground-truth* that includes
non-additive interactions (brand × container shared cleaning, line-specific
penalties, family-conflict bumps, packaging-direction asymmetry).

* :func:`_ground_truth_segments` — synthesises the "real plant" behaviour
  (private). Stored in the Excel for inspection.
* :func:`get_transition_cost`     — LightGBM **q50** prediction.
* :func:`get_transition_uncertainty` — ``(q10, q90)`` band from a second pair
  of LightGBM models trained with quantile objective.
* :func:`get_transition_segments` — ground-truth decomposition (kept for
  Excel inspectability; sequencer and partitioner consume :func:`get_transition_cost`).
* :func:`get_node_cost`           — production hours per line (unchanged).

The models are trained once on a 10k-sample synthetic history and cached to
``data_experiment/changeover_model.pkl``. Re-train by deleting the cache.

When real Damm history lands, replace the body of :func:`_train_models`
with a load of the actual ``Cambios`` table — the rest of the pipeline is
untouched.
"""

from __future__ import annotations

import pickle
import random
from dataclasses import asdict, dataclass, fields
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import lightgbm as lgb
import numpy as np
from openpyxl import Workbook, load_workbook

# ---------------------------------------------------------------------------
# Determinism — single seed governs the whole synthetic dataset
# ---------------------------------------------------------------------------
SEED: int = 42

# ---------------------------------------------------------------------------
# Domain constants (mirror the LineWise hard capability matrix)
# ---------------------------------------------------------------------------

LINE_IDS: tuple[int, ...] = (14, 17, 19)

LINE_CONTAINER_TYPES: dict[int, frozenset[str]] = {
    14: frozenset({"1/3", "1/2"}),
    17: frozenset({"1/3"}),
    19: frozenset({"1/3", "1/2", "2/5"}),
}

CONTAINER_TYPE_CL: dict[str, int] = {"1/3": 33, "1/2": 50, "2/5": 44}

BRANDS: tuple[str, ...] = (
    "ESTRELLA", "VOLL_DAMM", "FREE_DAMM", "DAURA", "INEDIT", "TURIA",
)
FAMILIES: tuple[str, ...] = (
    "LAGER", "PREMIUM", "SIN_GLUTEN", "SIN_ALCOHOL", "RADLER",
)
PRIMARY_PACKS: tuple[str, ...] = ("P6", "P12", "P24")
SECONDARY_PACKS: tuple[str, ...] = ("BANDEJA", "CAJA", "PALET_DIR")

_PACK_ORDER: dict[str, int] = {p: i for i, p in enumerate(PRIMARY_PACKS)}

# ---------------------------------------------------------------------------
# SKU metadata
# ---------------------------------------------------------------------------

@dataclass(frozen=True)
class SkuMeta:
    sku_id: str
    container_type: str
    brand: str
    family: str
    primary_packaging: str
    secondary_packaging: str
    units_per_case: int
    volume_cl: int


def build_sku_catalog(
    n_per_container_type: dict[str, int] | None = None,
    *,
    rng: random.Random | None = None,
) -> list[SkuMeta]:
    n_per_container_type = n_per_container_type or {"1/3": 15, "1/2": 10, "2/5": 5}
    rng = rng or random.Random(SEED)
    catalog: list[SkuMeta] = []
    for ct, n in n_per_container_type.items():
        for i in range(n):
            brand = rng.choice(BRANDS)
            pp = rng.choice(PRIMARY_PACKS)
            sku = SkuMeta(
                sku_id=f"{brand[:3]}_{CONTAINER_TYPE_CL[ct]}_{pp}_{i:02d}",
                container_type=ct,
                brand=brand,
                family=rng.choice(FAMILIES),
                primary_packaging=pp,
                secondary_packaging=rng.choice(SECONDARY_PACKS),
                units_per_case={"P6": 6, "P12": 12, "P24": 24}[pp],
                volume_cl=CONTAINER_TYPE_CL[ct],
            )
            catalog.append(sku)
    return catalog


DEMO_WINDOW_ID: str = "2026-W21-7d"
DEMO_WINDOW_START: date = date(2026, 5, 18)
DEMO_WINDOW_END: date = DEMO_WINDOW_START + timedelta(days=6)


def build_demand(
    catalog: list[SkuMeta],
    *,
    window_id: str = DEMO_WINDOW_ID,
    window_start: date = DEMO_WINDOW_START,
    window_end: date = DEMO_WINDOW_END,
    rng: random.Random | None = None,
) -> list[dict[str, Any]]:
    rng = rng or random.Random(SEED + 1)
    return [
        {
            "window_id": window_id,
            "window_start": window_start.isoformat(),
            "window_end": window_end.isoformat(),
            "sku_id": sku.sku_id,
            "units_demanded": rng.randint(50_000, 500_000),
            "source": "whatif_usuario",
            "priority": rng.randint(2, 4),
        }
        for sku in catalog
    ]


# ---------------------------------------------------------------------------
# === COST MODEL — ground truth + LightGBM predictor ◀──── SWAP TARGET ===
# ---------------------------------------------------------------------------

LINE_SPEED_MOD: dict[int, float] = {14: 1.00, 17: 1.05, 19: 0.95}
BASE_SPEED_UDS_H: int = 60_000
RAMP_UP_HOURS: float = 0.5

# Per-segment additive weights (the *baseline* — ground truth then perturbs
# these via interactions and noise).
SEGMENT_WEIGHTS: dict[str, float] = {
    "container":      3.0,
    "brand":          1.0,
    "family":         0.5,
    "primary_pack":   0.5,
    "secondary_pack": 0.25,
    "base":           0.25,
}

# Ground-truth interaction multipliers — the non-additive structure the ML
# must learn. Documented so reviewers can see what the model is up against.
_GT_INTERACTIONS = {
    # When BOTH brand and container change, cleaning can be combined → save 25%
    # on container time. Classic SMED finding.
    "brand_container_shared_clean": 0.75,
    # SIN_GLUTEN downstream needs extra sterilisation regardless of upstream.
    "sin_gluten_destination_penalty_h": 0.80,
    # L17 has older brand-change tooling → 30% slower on brand swaps.
    "l17_brand_penalty": 1.30,
    # L19 has automatic container-change carrousel → 20% faster on container.
    "l19_container_bonus": 0.80,
    # Packaging direction: going to a smaller pack (P24 → P6) is faster than
    # the reverse (re-tooling the case packer).
    "pack_downsize_bonus": 0.85,
    "pack_upsize_penalty": 1.20,
    # Gaussian observation noise (std in hours) — simulates measurement error.
    "noise_std_h": 0.15,
}


def _ground_truth_segments(
    sku_a: SkuMeta, sku_b: SkuMeta, line_id: int, *, rng: random.Random | None = None,
) -> dict[str, float]:
    """The hidden 'real plant' cost function — non-additive, line-dependent.

    Used to (a) train the LightGBM predictor and (b) populate the
    ``edge_matrix`` segment columns so reviewers can compare prediction vs
    reality. Sequencer / partitioner consume :func:`get_transition_cost`
    (ML prediction), not this function.
    """
    seg: dict[str, float] = {"base": SEGMENT_WEIGHTS["base"]}
    diff_container = sku_a.container_type != sku_b.container_type
    diff_brand = sku_a.brand != sku_b.brand
    diff_family = sku_a.family != sku_b.family
    diff_pp = sku_a.primary_packaging != sku_b.primary_packaging
    diff_sp = sku_a.secondary_packaging != sku_b.secondary_packaging

    if diff_container:
        c = SEGMENT_WEIGHTS["container"]
        if diff_brand:
            c *= _GT_INTERACTIONS["brand_container_shared_clean"]
        if line_id == 19:
            c *= _GT_INTERACTIONS["l19_container_bonus"]
        seg["container"] = c

    if diff_brand:
        b = SEGMENT_WEIGHTS["brand"]
        if line_id == 17:
            b *= _GT_INTERACTIONS["l17_brand_penalty"]
        seg["brand"] = b

    if diff_family:
        f = SEGMENT_WEIGHTS["family"]
        seg["family"] = f
    # Sin-gluten downstream is a destination-driven penalty, not symmetric.
    if sku_b.family == "SIN_GLUTEN" and sku_a.family != "SIN_GLUTEN":
        seg["family"] = seg.get("family", 0.0) + _GT_INTERACTIONS[
            "sin_gluten_destination_penalty_h"
        ]

    if diff_pp:
        p = SEGMENT_WEIGHTS["primary_pack"]
        d = _PACK_ORDER[sku_b.primary_packaging] - _PACK_ORDER[sku_a.primary_packaging]
        if d < 0:
            p *= _GT_INTERACTIONS["pack_downsize_bonus"]
        elif d > 0:
            p *= _GT_INTERACTIONS["pack_upsize_penalty"]
        seg["primary_pack"] = p

    if diff_sp:
        seg["secondary_pack"] = SEGMENT_WEIGHTS["secondary_pack"]

    if rng is not None:
        # Per-segment Gaussian noise — bigger segments get bigger absolute
        # noise (multiplicative on the segment, additive on the total).
        noise_total = rng.gauss(0.0, _GT_INTERACTIONS["noise_std_h"])
        seg["base"] = max(0.05, seg["base"] + noise_total)
    return {k: round(v, 4) for k, v in seg.items()}


def get_transition_segments(sku_a: SkuMeta, sku_b: SkuMeta, line_id: int = 14) -> dict[str, float]:
    """Public segment decomposition — uses ground truth (no noise).

    Sequencer / partitioner do NOT call this; they call
    :func:`get_transition_cost` which goes through the ML model. This
    function exists for human inspectability via the ``edge_matrix`` sheet.
    """
    return _ground_truth_segments(sku_a, sku_b, line_id, rng=None)


# --- Featurisation ---------------------------------------------------------

# Feature schema for the LightGBM models — kept tiny and interpretable so the
# trained boosters are SHAP-friendly. Categorical features are one-hot via
# integer-encoded brand/family/pp to keep the example dependency-light.
_BRAND_IX = {b: i for i, b in enumerate(BRANDS)}
_FAMILY_IX = {f: i for i, f in enumerate(FAMILIES)}
_CONTAINER_IX = {c: i for i, c in enumerate(CONTAINER_TYPE_CL.keys())}
_LINE_IX = {l: i for i, l in enumerate(LINE_IDS)}

_FEATURE_NAMES: tuple[str, ...] = (
    "diff_container", "diff_brand", "diff_family", "diff_pp", "diff_sp",
    "line_ix", "brand_a_ix", "brand_b_ix",
    "family_a_ix", "family_b_ix",
    "container_a_ix", "container_b_ix",
    "pp_delta",
    "brand_x_container", "sin_gluten_dest", "l17_x_brand", "l19_x_container",
)


def _featurise(sku_a: SkuMeta, sku_b: SkuMeta, line_id: int) -> list[float]:
    """Translate a (sku_a, sku_b, line) triple into the model's feature row."""
    diff_container = float(sku_a.container_type != sku_b.container_type)
    diff_brand = float(sku_a.brand != sku_b.brand)
    diff_family = float(sku_a.family != sku_b.family)
    diff_pp = float(sku_a.primary_packaging != sku_b.primary_packaging)
    diff_sp = float(sku_a.secondary_packaging != sku_b.secondary_packaging)
    pp_delta = float(
        _PACK_ORDER[sku_b.primary_packaging] - _PACK_ORDER[sku_a.primary_packaging]
    )
    return [
        diff_container, diff_brand, diff_family, diff_pp, diff_sp,
        float(_LINE_IX[line_id]),
        float(_BRAND_IX[sku_a.brand]), float(_BRAND_IX[sku_b.brand]),
        float(_FAMILY_IX[sku_a.family]), float(_FAMILY_IX[sku_b.family]),
        float(_CONTAINER_IX[sku_a.container_type]),
        float(_CONTAINER_IX[sku_b.container_type]),
        pp_delta,
        diff_brand * diff_container,
        float(sku_b.family == "SIN_GLUTEN" and sku_a.family != "SIN_GLUTEN"),
        float(line_id == 17) * diff_brand,
        float(line_id == 19) * diff_container,
    ]


# --- LightGBM training (cached) -------------------------------------------

_MODEL_CACHE_PATH: Path = (
    Path(__file__).resolve().parent / "data_experiment" / "changeover_model.pkl"
)
_MODELS: dict[str, lgb.Booster] | None = None


def _sample_training_set(
    n_samples: int, *, seed: int,
) -> tuple[np.ndarray, np.ndarray]:
    """Draw (X, y) from the ground truth — n_samples random feasible triples."""
    rng = random.Random(seed)
    catalog = build_sku_catalog(rng=random.Random(seed + 100))
    X_rows: list[list[float]] = []
    y: list[float] = []
    while len(y) < n_samples:
        a = rng.choice(catalog)
        b = rng.choice(catalog)
        if a.sku_id == b.sku_id:
            continue
        line_id = rng.choice(LINE_IDS)
        if a.container_type not in LINE_CONTAINER_TYPES[line_id]:
            continue
        if b.container_type not in LINE_CONTAINER_TYPES[line_id]:
            continue
        segs = _ground_truth_segments(a, b, line_id, rng=rng)
        X_rows.append(_featurise(a, b, line_id))
        y.append(sum(segs.values()))
    return np.asarray(X_rows, dtype=np.float64), np.asarray(y, dtype=np.float64)


def _train_models(n_samples: int = 10_000, *, seed: int = SEED) -> dict[str, lgb.Booster]:
    """Train three LightGBM quantile regressors (q10, q50, q90)."""
    X, y = _sample_training_set(n_samples, seed=seed)
    n_train = int(0.85 * len(y))
    X_tr, X_va = X[:n_train], X[n_train:]
    y_tr, y_va = y[:n_train], y[n_train:]

    common = dict(
        objective="quantile",
        learning_rate=0.05,
        num_leaves=31,
        min_data_in_leaf=20,
        feature_fraction=0.9,
        bagging_fraction=0.9,
        bagging_freq=5,
        verbose=-1,
        seed=seed,
    )
    models: dict[str, lgb.Booster] = {}
    for tag, alpha in (("q10", 0.10), ("q50", 0.50), ("q90", 0.90)):
        params = dict(common, alpha=alpha)
        dtr = lgb.Dataset(X_tr, label=y_tr, feature_name=list(_FEATURE_NAMES))
        dva = lgb.Dataset(X_va, label=y_va, feature_name=list(_FEATURE_NAMES))
        booster = lgb.train(
            params, dtr, num_boost_round=400, valid_sets=[dva],
            callbacks=[lgb.early_stopping(stopping_rounds=30, verbose=False)],
        )
        models[tag] = booster
    return models


def _ensure_models() -> dict[str, lgb.Booster]:
    """Load cached models or train + cache them."""
    global _MODELS
    if _MODELS is not None:
        return _MODELS
    if _MODEL_CACHE_PATH.exists():
        with _MODEL_CACHE_PATH.open("rb") as fh:
            payload = pickle.load(fh)
        _MODELS = {tag: lgb.Booster(model_str=s) for tag, s in payload.items()}
        return _MODELS
    _MODEL_CACHE_PATH.parent.mkdir(parents=True, exist_ok=True)
    models = _train_models()
    with _MODEL_CACHE_PATH.open("wb") as fh:
        pickle.dump({tag: m.model_to_string() for tag, m in models.items()}, fh)
    _MODELS = models
    return _MODELS


def _predict(tag: str, sku_a: SkuMeta, sku_b: SkuMeta, line_id: int) -> float:
    booster = _ensure_models()[tag]
    x = np.asarray([_featurise(sku_a, sku_b, line_id)], dtype=np.float64)
    pred = booster.predict(x)[0]
    return max(0.1, float(pred))


def get_transition_cost(
    sku_a: SkuMeta, sku_b: SkuMeta, line_id: int = 14,
) -> float:
    """LightGBM **q50** prediction of changeover hours.

    This is what the sequencer and partitioner consume — i.e. they optimise
    against the *model's belief* about the plant, not the ground truth.
    The optimiser is therefore as good as the model.
    """
    return round(_predict("q50", sku_a, sku_b, line_id), 4)


def get_transition_uncertainty(
    sku_a: SkuMeta, sku_b: SkuMeta, line_id: int = 14,
) -> tuple[float, float]:
    """Prediction band ``(q10, q90)`` from the quantile ensemble.

    Useful for robust optimisation extensions (penalise edges with wide
    bands, surface "uncertain transitions" in the UI). The current
    partitioner ignores it but the contract is in place.
    """
    q10 = _predict("q10", sku_a, sku_b, line_id)
    q90 = _predict("q90", sku_a, sku_b, line_id)
    if q90 < q10:
        q10, q90 = q90, q10
    return round(q10, 4), round(q90, 4)


def get_node_cost(sku: SkuMeta, units_demanded: int, line_id: int) -> float:
    speed = BASE_SPEED_UDS_H * LINE_SPEED_MOD[line_id]
    return round(units_demanded / speed + RAMP_UP_HOURS, 4)


def median_speed_uds_per_hour(line_id: int) -> float:
    return float(BASE_SPEED_UDS_H * LINE_SPEED_MOD[line_id])


# ---------------------------------------------------------------------------
# Materialisation
# ---------------------------------------------------------------------------

ALL_SEGMENTS: tuple[str, ...] = tuple(SEGMENT_WEIGHTS.keys())


def materialise_line_capability(catalog: list[SkuMeta]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sku in catalog:
        for line_id in LINE_IDS:
            can = sku.container_type in LINE_CONTAINER_TYPES[line_id]
            rows.append({
                "sku_id": sku.sku_id,
                "line_id": line_id,
                "can_produce": bool(can),
                "median_speed_uds_per_hour": median_speed_uds_per_hour(line_id),
            })
    return rows


def materialise_edge_matrix(catalog: list[SkuMeta]) -> list[dict[str, Any]]:
    """Edge rows with ground-truth segments + LightGBM prediction band."""
    rows: list[dict[str, Any]] = []
    for line_id in LINE_IDS:
        allowed = LINE_CONTAINER_TYPES[line_id]
        feasible = [s for s in catalog if s.container_type in allowed]
        for a in feasible:
            for b in feasible:
                if a.sku_id == b.sku_id:
                    continue
                segments = get_transition_segments(a, b, line_id)
                gt_total = round(sum(segments.values()), 4)
                pred = get_transition_cost(a, b, line_id)
                q10, q90 = get_transition_uncertainty(a, b, line_id)
                row: dict[str, Any] = {
                    "line_id": line_id,
                    "sku_from_id": a.sku_id,
                    "sku_to_id": b.sku_id,
                    "total_hours": gt_total,
                    "predicted_hours": pred,
                    "q10_hours": q10,
                    "q90_hours": q90,
                    "source": "ml",
                }
                for seg_name in ALL_SEGMENTS:
                    row[f"{seg_name}_h"] = round(segments.get(seg_name, 0.0), 4)
                rows.append(row)
    return rows


def materialise_node_costs(
    catalog: list[SkuMeta], demand: list[dict[str, Any]],
) -> list[dict[str, Any]]:
    units_by_sku = {row["sku_id"]: row["units_demanded"] for row in demand}
    rows: list[dict[str, Any]] = []
    for sku in catalog:
        units = int(units_by_sku.get(sku.sku_id, 0))
        for line_id in LINE_IDS:
            if sku.container_type not in LINE_CONTAINER_TYPES[line_id]:
                continue
            rows.append({
                "line_id": line_id,
                "sku_id": sku.sku_id,
                "units_demanded": units,
                "production_hours": get_node_cost(sku, units, line_id),
            })
    return rows


# ---------------------------------------------------------------------------
# Excel I/O
# ---------------------------------------------------------------------------

def _write_sheet(wb: Workbook, sheet_name: str, rows: list[dict[str, Any]]) -> None:
    ws = wb.create_sheet(sheet_name)
    if not rows:
        return
    headers = list(rows[0].keys())
    ws.append(headers)
    for row in rows:
        ws.append([row.get(h) for h in headers])


def _write_catalog_sheet(wb: Workbook, sheet_name: str, catalog: list[SkuMeta]) -> None:
    ws = wb.create_sheet(sheet_name)
    headers = [f.name for f in fields(SkuMeta)]
    ws.append(headers)
    for sku in catalog:
        d = asdict(sku)
        ws.append([d[h] for h in headers])


def read_sheet(path: Path, sheet_name: str) -> list[dict[str, Any]]:
    wb = load_workbook(path, data_only=True, read_only=True)
    try:
        ws = wb[sheet_name]
        rows_iter = ws.iter_rows(values_only=True)
        try:
            headers = list(next(rows_iter))
        except StopIteration:
            return []
        return [dict(zip(headers, row)) for row in rows_iter]
    finally:
        wb.close()


# ---------------------------------------------------------------------------
# Validation helpers (for the demo / pitch)
# ---------------------------------------------------------------------------

def evaluate_model_accuracy(n_samples: int = 2_000, *, seed: int = SEED + 7) -> dict[str, float]:
    """Quick MAE / coverage check of the LightGBM predictor vs ground truth."""
    X, y = _sample_training_set(n_samples, seed=seed)
    models = _ensure_models()
    pred = models["q50"].predict(X)
    q10 = models["q10"].predict(X)
    q90 = models["q90"].predict(X)
    abs_err = np.abs(pred - y)
    coverage = float(np.mean((y >= q10) & (y <= q90)))
    return {
        "mae_hours": float(np.mean(abs_err)),
        "rmse_hours": float(np.sqrt(np.mean(abs_err ** 2))),
        "p50_band_hours": float(np.median(q90 - q10)),
        "coverage_80pct": coverage,  # target: ~0.80
        "n_samples": float(n_samples),
    }


# ---------------------------------------------------------------------------
# Entry point
# ---------------------------------------------------------------------------

DEFAULT_OUT_PATH: Path = (
    Path(__file__).resolve().parent / "data_experiment" / "artificial_plans.xlsx"
)


def write_excel(out_path: Path = DEFAULT_OUT_PATH) -> Path:
    catalog = build_sku_catalog()
    demand = build_demand(catalog)
    capability = materialise_line_capability(catalog)
    nodes = materialise_node_costs(catalog, demand)
    _ensure_models()  # train once before the edge loop touches the predictor
    edges = materialise_edge_matrix(catalog)

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    default = wb.active
    wb.remove(default)
    _write_catalog_sheet(wb, "skus", catalog)
    _write_sheet(wb, "demand", demand)
    _write_sheet(wb, "line_capability", capability)
    _write_sheet(wb, "node_costs", nodes)
    _write_sheet(wb, "edge_matrix", edges)
    wb.save(out_path)
    return out_path


if __name__ == "__main__":
    catalog = build_sku_catalog()
    print("[generate_test_data] ensuring LightGBM models...")
    _ensure_models()
    print("[generate_test_data] writing Excel...")
    written = write_excel()
    metrics = evaluate_model_accuracy()
    print(f"[generate_test_data] wrote {written}")
    print(
        f"  skus            : {len(catalog)} rows\n"
        f"  line_capability : {len(materialise_line_capability(catalog))} rows\n"
        f"  edge_matrix     : {len(materialise_edge_matrix(catalog))} rows"
    )
    print(
        "\n  changeover model (LightGBM quantile ensemble):"
        f"\n    MAE          = {metrics['mae_hours']:.3f} h"
        f"\n    RMSE         = {metrics['rmse_hours']:.3f} h"
        f"\n    P50 band     = {metrics['p50_band_hours']:.3f} h"
        f"\n    80%% coverage = {metrics['coverage_80pct']*100:.1f}%%"
    )
